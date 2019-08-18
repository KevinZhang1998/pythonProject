import openpyxl
from collections import Counter
from selenium import webdriver

def countItem(li):
    result = {}
    for i in set(li):
        result[i] = li.count(i)
    return result

if __name__ == '__main__':
    wb = openpyxl.load_workbook("example.xlsx")
    ws = wb.active

    colA = ws['D']
    rowRange = ws[1:20]
    ele = []
    subClassNumber = {}
    countNumber = 1
    for cell in colA:
        ele.append(cell.value)
    result = countItem(ele)
    print(result)


